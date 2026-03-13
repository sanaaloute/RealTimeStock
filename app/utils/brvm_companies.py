"""Load BRVM companies and symbols from data/BRVM_Companies.xlsx for NLU mapping (avoid hallucination)."""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Path relative to app package (app/data/BRVM_Companies.xlsx)
BRVM_COMPANIES_FILE = Path(__file__).resolve().parent.parent / "data" / "BRVM_Companies.xlsx"

# Excel column names: Name, Symbol, Sector, industryCluster, Country
COL_NAME = "Name"
COL_SYMBOL = "Symbol"
COL_SECTOR = "Sector"
COL_INDUSTRY_CLUSTER = "industryCluster"
COL_COUNTRY = "Country"

_symbol_to_name: dict[str, str] = {}
_symbol_to_sector: dict[str, str] = {}
_symbol_to_country: dict[str, str] = {}  # symbol -> 2-letter country code (ml, sn, ci, bj, ...)
_name_to_symbol: dict[str, str] = {}
_valid_symbols: set[str] = set()
_loaded = False

# BRVM country name (from Excel) -> Sika Finance URL suffix (2-letter)
_COUNTRY_TO_CODE: dict[str, str] = {
    "mali": "ml", "sénégal": "sn", "senegal": "sn", "côte d'ivoire": "ci", "cote d'ivoire": "ci",
    "bénin": "bj", "benin": "bj", "burkina faso": "bf", "niger": "ne", "togo": "tg",
    "guinée-bissau": "gw", "guinee-bissau": "gw",
}

# Common words that cause wrong symbol resolution when mapped by single word
_STOP_WORDS = frozenset({
    "bank", "africa", "côte", "d'ivoire", "société", "de", "du", "des", "ci", "internationale",
    "international", "bénin", "burkina", "faso", "mali", "niger", "sénégal", "togo",
})


def _normalize(s: str) -> str:
    """Lowercase, strip, collapse spaces for lookup."""
    return " ".join((s or "").lower().strip().split())


def _cell_value(v: Any) -> str:
    """Get string value from an openpyxl Cell or a raw value."""
    if v is None:
        return ""
    if hasattr(v, "value"):
        v = getattr(v, "value")
    if v is None:
        return ""
    return str(v).strip()


def _load() -> None:
    global _symbol_to_name, _symbol_to_sector, _symbol_to_country, _name_to_symbol, _valid_symbols, _loaded
    if _loaded:
        return
    _symbol_to_name = {}
    _symbol_to_sector = {}
    _symbol_to_country = {}
    _name_to_symbol = {}
    _valid_symbols = set()
    if not BRVM_COMPANIES_FILE.exists():
        logger.warning("BRVM companies file not found: %s", BRVM_COMPANIES_FILE)
        _loaded = True
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed. pip install openpyxl to load BRVM_Companies.xlsx")
        _loaded = True
        return
    try:
        wb = load_workbook(BRVM_COMPANIES_FILE, read_only=True, data_only=True)
        sheet = wb.active
        if sheet is None:
            wb.close()
            _loaded = True
            return
        rows = list(sheet.iter_rows(min_row=1, values_only=False))
        wb.close()
    except Exception as e:
        logger.warning("Failed to load BRVM companies from %s: %s", BRVM_COMPANIES_FILE, e)
        _loaded = True
        return
    if not rows:
        _loaded = True
        return
    # First row: header; match column names (case-insensitive for robustness)
    header = [_cell_value(c) for c in rows[0]]
    col_index: dict[str, int] = {}
    for i, h in enumerate(header):
        key = (h or "").strip()
        if key:
            col_index[key.lower()] = i
    name_col = col_index.get(COL_NAME.lower(), col_index.get("name", -1))
    symbol_col = col_index.get(COL_SYMBOL.lower(), col_index.get("symbol", -1))
    sector_col = col_index.get(COL_SECTOR.lower(), col_index.get("sector", -1))
    industry_col = col_index.get(COL_INDUSTRY_CLUSTER.lower(), -1)
    country_col = col_index.get(COL_COUNTRY.lower(), col_index.get("country", -1))
    if symbol_col < 0 or name_col < 0:
        logger.warning("BRVM_Companies.xlsx: expected columns %r and %r in %s", COL_SYMBOL, COL_NAME, list(col_index))
        _loaded = True
        return
    for row in rows[1:]:
        if not row:
            continue
        cells = list(row)
        symbol = _cell_value(cells[symbol_col]) if symbol_col < len(cells) else ""
        name_part = _cell_value(cells[name_col]) if name_col < len(cells) else ""
        if not symbol:
            continue
        symbol = symbol.upper()
        sector = _cell_value(cells[sector_col]) if sector_col >= 0 and sector_col < len(cells) else ""
        if not sector and industry_col >= 0 and industry_col < len(cells):
            sector = _cell_value(cells[industry_col])
        country_name = _cell_value(cells[country_col]) if country_col >= 0 and country_col < len(cells) else ""
        _valid_symbols.add(symbol)
        _symbol_to_name[symbol] = name_part or symbol
        if sector:
            _symbol_to_sector[symbol] = sector
        if country_name:
            code = _COUNTRY_TO_CODE.get(country_name.strip().lower(), "").strip() or country_name.strip()[:2].lower() if len(country_name.strip()) >= 2 else ""
            if code:
                _symbol_to_country[symbol] = code
        _name_to_symbol[_normalize(name_part)] = symbol
        _name_to_symbol[_normalize(symbol)] = symbol
        for word in (name_part or "").replace(",", " ").replace("|", " ").replace("(", " ").replace(")", " ").split():
            w = word.strip()
            if len(w) >= 3 and w.upper() != symbol and _normalize(w) not in _STOP_WORDS:
                _name_to_symbol[_normalize(w)] = symbol
    _loaded = True


def get_valid_symbols() -> set[str]:
    """Return set of valid BRVM symbols."""
    _load()
    return set(_valid_symbols)


def get_symbol_to_name() -> dict[str, str]:
    """Return mapping symbol -> company name."""
    _load()
    return dict(_symbol_to_name)


def get_symbol_to_sector() -> dict[str, str]:
    """Return mapping symbol -> sector (empty string if not set)."""
    _load()
    return dict(_symbol_to_sector)


def get_symbol_to_country() -> dict[str, str]:
    """Return mapping symbol -> 2-letter country code (ml, sn, ci, bj, etc.) for Sika Finance URLs."""
    _load()
    return dict(_symbol_to_country)


def get_country_code_for_symbol(symbol: str) -> str:
    """Return 2-letter country code for symbol. Default 'ci' (Côte d'Ivoire) if unknown."""
    _load()
    sym = (symbol or "").strip().upper()
    return _symbol_to_country.get(sym, "ci")


def get_name_to_symbol() -> dict[str, str]:
    """Return mapping normalized name -> symbol."""
    _load()
    return dict(_name_to_symbol)


# Common abbreviations (user may type "SMB" for SMBC, etc.)
_ABBREVS: dict[str, str] = {
    "smb": "SMBC",  # SMB Société Minière
}


def resolve_to_symbol(mention: str) -> str | None:
    """
    Resolve user mention (symbol or company name) to official symbol.
    Returns symbol if found, None if unknown (caller can ask for clarification).
    """
    _load()
    mention = (mention or "").strip()
    if not mention:
        return None
    upper = mention.upper()
    if upper in _valid_symbols:
        return upper
    norm = _normalize(mention)
    if norm in _ABBREVS:
        return _ABBREVS[norm]
    by_name = _name_to_symbol.get(norm)
    if by_name:
        return by_name
    return None


def format_list_for_prompt() -> str:
    """Format the BRVM list for inclusion in NLU system prompt (symbol -> name [sector], one per line)."""
    _load()
    if not _symbol_to_name:
        return "No BRVM company list loaded."
    lines = []
    for sym, name in sorted(_symbol_to_name.items()):
        sector = _symbol_to_sector.get(sym, "")
        if sector:
            lines.append(f"- {sym}: {name} ({sector})")
        else:
            lines.append(f"- {sym}: {name}")
    return "\n".join(lines)


def normalize_entities(entities: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    """
    Resolve symbol, symbol_a, symbol_b from entities using the BRVM list.
    Returns (updated_entities, list of unknown mentions to report).
    """
    _load()
    out = dict(entities)
    unknown: list[str] = []
    for key in ("symbol", "symbol_a", "symbol_b"):
        val = out.get(key)
        if not val or not isinstance(val, str):
            continue
        resolved = resolve_to_symbol(val)
        if resolved:
            out[key] = resolved
        else:
            unknown.append(val)
    return out, unknown
